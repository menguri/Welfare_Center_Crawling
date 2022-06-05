from msilib.schema import Class
import re
import openpyxl
from selenium.webdriver.chrome.options import Options
import selenium
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from time import sleep
import datetime
import pandas as pd
from urllib.request import urlopen
import requests
from bs4 import BeautifulSoup
from html_table_parser import parser_functions as parser
import os
import sys
# python 3.9
import pandas as pd
import warnings
warnings.simplefilter("ignore")
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.font_manager as fm
import requests
import json
import folium

# 폰트 설정하기 ---------------------------------------------------------------------------------------------------------------------------------------------------------
def get_font():
    # ttf 폰트 전체갯수
    f_list = [(f.name, f.fname) for f in fm.fontManager.ttflist if "Malgun Gothic" in f.name]
    # fname 옵션을 사용하는 방법
    path = f_list[0][1]
    font_name = fm.FontProperties(fname=path, size=18).get_name()
    plt.rc('font', family=font_name)


# 위/경도 변환 함수 -----------------------------------------------------------------------------------------------------------------------------------------------------
from geopy.geocoders import Nominatim
geo_local = Nominatim(user_agent='South Korea')
# 위도, 경도 반환하는 함수
def geocoding(address):
    geo = geo_local.geocode(address)
    x_y = [geo.latitude, geo.longitude]
    return x_y

def lat_long(df):
    # week 주소를 위, 경도 값으로 변환하기
    latitude = []
    longitude =[]

    for i in df['주소']:
        try:
            lati_longi = geocoding(i)
            latitude.append(lati_longi[0])
            longitude.append(lati_longi[1])
        except:
            latitude.append(36.082402)
            longitude.append(128.2715050)
    df['위도'] = latitude
    df['경도'] = longitude
    return df

# 전처리 날짜 확인 -----------------------------------------------------------------------------------------------------------------------------------------------------
import datetime
df_now = datetime.datetime.now()

# df to pdf ------------------------------------------------------------------------------------------------------------------------------------------------------------
def df2excel(df, title, program_directory):
    df.to_excel(program_directory + '/' + f"{df_now.month}월{df_now.day}일" + '/' + f'{title}.xlsx')

# pandas option
pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', None)
pd.set_option('display.max_colwidth', -1)


# center file 전처리 -----------------------------------------------------------------------------------------------------------------------------------------------------
def centerfile_processing(program_directory):
    get_font()
    import datetime
    df_now = datetime.datetime.now()
    # 폴더 생성
    os.mkdir(program_directory + '/' + f"{df_now.month}월{df_now.day}일")
    # 센터 파일 불러오기
    center_file_path = program_directory + '/' + f"{df_now.month}월{df_now.day}일_센터 현황.xlsx"
    center_wb = openpyxl.load_workbook(center_file_path, data_only=True)
    # default 시트 열기
    ws = center_wb.active
    # A행, 1-4열 삭제
    ws.delete_rows(1, 4)
    ws.delete_cols(1,2)
    # dataframe으로 굽기
    df = pd.DataFrame(ws.values)
    df.columns = ['장기요양기관', '급여종류', '평가결과', '정원', '현원', '잔여', '대기', '방문목욕차량', '주소', '전화번호']
    # 주소/방문목욕차량 전처리
    df = df.drop('방문목욕차량', axis=1)
    df['주소'] = df.주소.apply(lambda x: ' '.join(x.split(' ')[0:4]))
    # 주야간보호센터와 방문요양센터 나누기
    week_center = df[df['급여종류']=='주야간보호'].copy()
    week_center = lat_long(week_center)
    df2excel(week_center, f"{df_now.month}월_{df_now.day}일_주야간보호 현황", program_directory)
    visit_center = df[df['급여종류']=='방문요양'].copy()
    visit_center = lat_long(visit_center)
    df2excel(visit_center, f"{df_now.month}월_{df_now.day}일_방문요양 현황", program_directory)
    # 지도에 표시해주기 -> 주야간은 파란색/ 방문요양센터는 빨간색, 주야간의 [현원/정원]은 동그라미에 바로 표시된다.
    center = [36.382402, 128.2715050]
    m = folium.Map(location=center, zoom_start=10)

    df_c = week_center[['장기요양기관', '정원', '현원', '위도', '경도']].copy()
    # 주간보호센터 데이터를 그려냅니다.
    for name, r_human, j_human, lat, lon in zip(df_c['장기요양기관'], df_c['정원'], df_c['현원'], df_c['위도'], df_c['경도']):
        folium.Marker(
            location = [lat, lon],
            popup= f"{name} : {j_human}/{r_human}",
            icon=folium.Icon(color='red',icon='star')
        ).add_to(m)

    df_h = visit_center[['장기요양기관','위도', '경도']].copy()
    # 방문요양센터 데이터를 그려냅니다.
    for name, lat, lon in zip(df_h['장기요양기관'], df_h['위도'], df_h['경도']):
        folium.Circle(
            location = [lat, lon],
            color = 'black'
        ).add_to(m)

    # geojson 사용할 수 있다면 사용하기
    m.save(program_directory + '/' + f"{df_now.month}월{df_now.day}일" + '/' + '센터 지도.html')


    # 그래프 그리기
    import seaborn as sns
    import matplotlib.pyplot as plt
    sns.set(font="Malgun Gothic", 
            rc={"axes.unicode_minus":False}, 
            style="darkgrid", 
            font_scale=4)

    # Initialize the matplotlib figure
    f, ax = plt.subplots(figsize=(100, 50))

    # Load the example car crash dataset
    df = df_c.sort_values("정원", ascending=False)
    df[['정원', '현원']] = df[['정원', '현원']].astype('int64')
    # Plot the total crashes
    sns.set_color_codes("pastel")
    sns.barplot(x="정원", y = "장기요양기관", data=df,
                label="정원", color="b")

    # Plot the crashes where alcohol was involved
    sns.set_color_codes("muted")
    sns.barplot(x="현원", y = "장기요양기관", data=df,
                label="현원", color="r")

    # Add a legend and informative axis label
    ax.legend(ncol=2, loc="lower right", frameon=True)
    ax.set(xlim=(0, 100), ylabel="인원 수", xlabel="장기요양기관")
    sns.despine(left=True, bottom=True)
    ax.set_title('장기요양기관 현원/정원')
    plt.savefig(program_directory + '/' + f"{df_now.month}월{df_now.day}일" + '/' + '장기요양기관 현원&정원.png')


    # 마무리: 파일 삭제
    if os.path.isfile(program_directory + '/' + f"{df_now.month}월{df_now.day}일_센터 현황.xlsx"):
        os.remove(program_directory + '/' + f"{df_now.month}월{df_now.day}일_센터 현황.xlsx")
        return 'okay'


# 상주 인구현황 file 전처리 -----------------------------------------------------------------------------------------------------------------------------------------------------
def humanfile_processing(program_directory):
    get_font()
    import datetime
    df_now = datetime.datetime.now()    
    human_file_path = program_directory + '/' + f"{df_now.month-1}월_상주인구현황.xlsx"
    human_wb = openpyxl.load_workbook(human_file_path, data_only=True)
    # default 시트 열기
    ws = human_wb.active

    # 칼럼 리스트 뽑기
    row = ws[6]
    big_list = []
    for i in row:
        big_list.append(i.value)

    big_list = list(filter(None, big_list))[1:]
    small_list = ['인구수', '구성비', '성비']

    # index 리스트 구성
    row = ws['B']
    row_list = []
    for i in row:
        row_list.append(i.value)
    row_list = list(filter(None, row_list))[4:]
    row_small_list = ['계', '남', '여'] 

    # 본격적인 df 구성
    ws.delete_rows(1,7)
    ws.delete_cols(1,3)
    df = pd.DataFrame(ws.values)

    # 칼럼 재구성
    columns_list = []
    for big in big_list:
        for small in small_list:
            columns_list.append(f'{big}_{small}')
    df.columns = columns_list
    for title in columns_list:
        if '구성비' in title or '성비' in title:
            df = df.drop(title, axis=1)
        else:
            df[title] = df[title].apply(lambda x: x.replace(',', ''))

    # index 재구성
    # [70:100]
    index_list = []
    for r in row_list:
        for r_m in row_small_list:
            index_list.append(f'{r}_{r_m}')

    df.index = index_list
    df = df[213:306]
    df = df.astype('int64')
    # 70-100세까지 남녀계 별로 합계만 내주기
    # '계' 
    df_sum = df[df.index.str.contains('계')]
    sum_df = pd.DataFrame(df_sum.sum(), columns=['합계']).transpose()
    # '남' 
    df_nam = df[df.index.str.contains('남')]
    nam_df = pd.DataFrame(df_nam.sum(), columns=['남']).transpose()
    # '여' 
    df_ye = df[df.index.str.contains('여')]
    ye_df = pd.DataFrame(df_ye.sum(), columns=['여']).transpose()

    df = pd.concat([sum_df, nam_df])
    df = pd.concat([df, ye_df])
    df = df.transpose()
    df.to_excel(program_directory + '/' + f"{df_now.month}월{df_now.day}일" + '/' + f"{df_now.month-1}월상주인구현황.xlsx")

    # 마무리: 파일 삭제
    if os.path.isfile(program_directory + '/' + f"{df_now.month-1}월_상주인구현황.xlsx"):
        os.remove(program_directory + '/' + f"{df_now.month-1}월_상주인구현황.xlsx")
        return 'okay'